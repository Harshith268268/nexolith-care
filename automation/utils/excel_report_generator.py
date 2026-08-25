"""
Excel Report Generator Utility.
Generates multi-sheet Excel workbooks with detailed formatting, metrics, pass/fail breakdowns, and defect summaries.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from automation.config.config import Config
from automation.utils.logger import AutomationLogger

from openpyxl.worksheet.worksheet import Worksheet

logger = AutomationLogger.get_logger()

class ExcelReportGenerator:
    @staticmethod
    def generate_all_excel_reports(results: List[Dict[str, Any]], metrics: Dict[str, Any]):
        """
        Generates Automation_Test_Report.xlsx, Failed_Test_Cases.xlsx, Passed_Test_Cases.xlsx, and Summary_Report.xlsx.
        """
        os.makedirs(Config.EXCEL_REPORTS_DIR, exist_ok=True)

        passed_tests = [r for r in results if r.get("status") == "PASS"]
        failed_tests = [r for r in results if r.get("status") == "FAIL"]
        skipped_tests = [r for r in results if r.get("status") in ["SKIP", "SKIPPED", "BLOCKED"]]

        # 1. Main Workbook: Automation_Test_Report.xlsx
        wb_main = openpyxl.Workbook()
        wb_main.remove(wb_main.active) # Remove default sheet

        # Sheet 1: Executed Test Cases
        ExcelReportGenerator._create_test_cases_sheet(wb_main, "Executed Test Cases", results)
        # Sheet 2: Passed Tests
        ExcelReportGenerator._create_test_cases_sheet(wb_main, "Passed Tests", passed_tests)
        # Sheet 3: Failed Tests
        ExcelReportGenerator._create_test_cases_sheet(wb_main, "Failed Tests", failed_tests)
        # Sheet 4: Skipped Tests
        ExcelReportGenerator._create_test_cases_sheet(wb_main, "Skipped Tests", skipped_tests)
        # Sheet 5: Execution Metrics
        ExcelReportGenerator._create_metrics_sheet(wb_main, "Execution Metrics", metrics)
        # Sheet 6: Defect Summary
        ExcelReportGenerator._create_defect_summary_sheet(wb_main, "Defect Summary", failed_tests)

        main_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Automation_Test_Report.xlsx")
        wb_main.save(main_path)
        logger.info(f"Generated Excel Report: {main_path}")

        # 2. Failed_Test_Cases.xlsx
        wb_failed = openpyxl.Workbook()
        ws_f = wb_failed.active
        ws_f.title = "Failed Tests"
        ExcelReportGenerator._populate_sheet(ws_f, failed_tests)
        failed_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Failed_Test_Cases.xlsx")
        wb_failed.save(failed_path)

        # 3. Passed_Test_Cases.xlsx
        wb_passed = openpyxl.Workbook()
        ws_p = wb_passed.active
        ws_p.title = "Passed Tests"
        ExcelReportGenerator._populate_sheet(ws_p, passed_tests)
        passed_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Passed_Test_Cases.xlsx")
        wb_passed.save(passed_path)

        # 4. Summary_Report.xlsx
        wb_sum = openpyxl.Workbook()
        ws_s = wb_sum.active
        ws_s.title = "Executive Summary"
        ExcelReportGenerator._create_metrics_sheet(wb_sum, "Executive Summary", metrics)
        sum_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Summary_Report.xlsx")
        wb_sum.save(sum_path)

    @staticmethod
    def _create_test_cases_sheet(wb: openpyxl.Workbook, title: str, test_cases: List[Dict[str, Any]]):
        ws = wb.create_sheet(title=title)
        ExcelReportGenerator._populate_sheet(ws, test_cases)

    @staticmethod
    def _populate_sheet(ws: Worksheet, test_cases: List[Dict[str, Any]]):

        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Preconditions", "Failure Reason"]
        
        # Header formatting
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        ws.row_dimensions[1].height = 25

        # Row Data
        for row_idx, tc in enumerate(test_cases, start=2):
            status = tc.get("status", "SKIP")
            row_data = [
                tc.get("test_id", ""),
                tc.get("module", ""),
                tc.get("name", ""),
                status,
                f"{tc.get('duration', 0.05):.2f}",
                tc.get("priority", "P2"),
                tc.get("preconditions", ""),
                tc.get("failure_reason", "")
            ]
            ws.append(row_data)

            # Cell Formatting
            status_cell = ws.cell(row=row_idx, column=4)
            if status == "PASS":
                status_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                status_cell.font = Font(color="166534", bold=True)
            elif status == "FAIL":
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                status_cell.font = Font(color="991B1B", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                status_cell.font = Font(color="92400E", bold=True)

            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                c.alignment = center_align if col_idx in [1, 4, 5, 6] else left_align

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @staticmethod
    def _create_metrics_sheet(wb: openpyxl.Workbook, title: str, metrics: Dict[str, Any]):
        ws = wb.create_sheet(title=title) if title not in wb.sheetnames else wb[title]
        ws.append(["Metric", "Value"])

        metric_rows = [
            ["Target Deployment BASE_URL", metrics.get("base_url", Config.BASE_URL)],
            ["Execution Timestamp", metrics.get("timestamp", "")],
            ["Total Test Cases Executed", metrics.get("total", 0)],
            ["Passed Test Cases", metrics.get("passed", 0)],
            ["Failed Test Cases", metrics.get("failed", 0)],
            ["Skipped Test Cases", metrics.get("skipped", 0)],
            ["Pass Percentage (%)", f"{metrics.get('pass_rate', 0.0):.2f}%"],
            ["Total Execution Time (sec)", f"{metrics.get('duration', 0.0):.2f} s"],
            ["Pipeline Build Status", metrics.get("status", "PASS")]
        ]

        for row in metric_rows:
            ws.append(row)

        for row_idx in range(1, len(metric_rows) + 2):
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40

    @staticmethod
    def _create_defect_summary_sheet(wb: openpyxl.Workbook, title: str, failed_tests: List[Dict[str, Any]]):
        ws = wb.create_sheet(title=title)
        ws.append(["Defect ID", "Test Case ID", "Module", "Failure Description", "Screenshot File"])
        for idx, ft in enumerate(failed_tests, start=1):
            ws.append([
                f"DEFECT-{idx:03d}",
                ft.get("test_id", ""),
                ft.get("module", ""),
                ft.get("failure_reason", "Assertion Failure"),
                ft.get("screenshot", "")
            ])
